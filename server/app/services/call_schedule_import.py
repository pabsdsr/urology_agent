from __future__ import annotations

from datetime import date, datetime
from io import BytesIO, StringIO
import csv
from typing import Any, Dict, Iterable, List, Tuple


POD_KEYS = ("North Pod", "Central Pod", "South Pod")


def _normalize_pod(pod_raw: Any) -> str | None:
    if pod_raw is None:
        return None
    s = str(pod_raw).strip()
    if not s:
        return None
    s_norm = " ".join(s.lower().split())
    if s_norm in ("north pod", "north pod.", "north"):
        return "North Pod"
    if s_norm in ("central pod", "central pod.", "central"):
        return "Central Pod"
    if s_norm in ("south pod", "south pod.", "south"):
        return "South Pod"
    return None


def _parse_header_date(value: Any) -> str:
    """
    Parse a header cell into ISO date string (YYYY-MM-DD).

    For CSV, we accept a variety of common date formats and normalize them
    to YYYY-MM-DD. For XLSX, we also accept real Excel date/datetime cells.
    """
    if value is None:
        raise ValueError("Empty date header")

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    s = str(value).strip()
    if not s:
        raise ValueError("Empty date header")

    # Try a range of common string formats and normalize to ISO.
    candidates = [
        "%Y-%m-%d",   # 2026-04-05 (ISO)
        "%Y/%m/%d",   # 2026/04/05
        "%m/%d/%Y",   # 04/05/2026
        "%m/%d/%y",   # 04/05/26
        "%m-%d-%Y",   # 04-05-2026
        "%m-%d-%y",   # 04-05-26
        "%d/%m/%Y",   # 05/04/2026 (day first)
        "%d/%m/%y",   # 05/04/26
        "%d-%m-%Y",   # 05-04-2026
        "%d-%m-%y",   # 05-04-26
        "%b %d %Y",   # Apr 05 2026
        "%d %b %Y",   # 05 Apr 2026
        "%B %d %Y",   # April 05 2026
        "%d %B %Y",   # 05 April 2026
    ]
    for fmt in candidates:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            continue

    raise ValueError(
        f"Invalid date header '{s}'. Please use a standard date format (e.g. 2026-04-05, 4/5/2026, or Apr 5 2026)"
    )


def _split_entries(cell_value: str) -> List[str]:
    """
    Accept either a single entry per cell, or multiple separated by newlines / semicolons.
    """
    raw = (cell_value or "").strip()
    if not raw:
        return []

    parts: List[str] = []
    for chunk in raw.replace("\r\n", "\n").split("\n"):
        chunk = chunk.strip()
        if not chunk:
            continue
        for sub in chunk.split(";"):
            sub = sub.strip()
            if sub:
                parts.append(sub)
    return parts


def _parse_location_practitioner(entry: str) -> Tuple[str, str]:
    s = (entry or "").strip()
    if not s:
        return ("", "")
    if ":" not in s:
        # Allow location-only as a fallback; practitioner remains empty.
        return (s, "")
    left, right = s.split(":", 1)
    return (left.strip(), right.strip())


def _empty_day_mapping_for_dates(dates: Iterable[str]) -> Dict[str, Dict[str, List[Dict[str, str]]]]:
    out: Dict[str, Dict[str, List[Dict[str, str]]]] = {}
    for d in dates:
        out[d] = {k: [] for k in POD_KEYS}
    return out


def _try_parse_header_date(value: Any) -> str | None:
    """Parse a header cell as date; return ISO date string or None if invalid."""
    try:
        return _parse_header_date(value)
    except Exception:
        return None


def _find_header_row(rows: List[List[Any]]) -> int:
    """Return index of first row that has at least one cell parsing as a date."""
    for r, row in enumerate(rows):
        if not row:
            continue
        for cell in row:
            if _try_parse_header_date(cell) is not None:
                return r
    raise ValueError(
        "No header row with dates found"
    )


def _find_pod_column(rows: List[List[Any]], header_row_idx: int) -> int:
    """Return leftmost column index that contains pod names (North/Central/South Pod) in data rows."""
    data_rows = rows[header_row_idx + 1 :]
    if not data_rows:
        raise ValueError("No data rows found below the header")
    max_col = max(len(row) for row in data_rows)
    for col in range(max_col):
        for row in data_rows[:20]:
            if col < len(row) and _normalize_pod(row[col]) is not None:
                return col
    raise ValueError(
        "No pod column found"
    )


def _find_date_columns(header_row: List[Any], pod_col: int) -> List[Tuple[int, str]]:
    """Return list of (column_index, date_str) for columns to the right of pod_col that parse as dates."""
    result: List[Tuple[int, str]] = []
    for c in range(pod_col + 1, len(header_row)):
        val = header_row[c] if c < len(header_row) else None
        d = _try_parse_header_date(val)
        if d is not None:
            result.append((c, d))
    return result


def parse_call_schedule_upload(file_bytes: bytes, filename: str) -> Dict[str, Dict[str, List[Dict[str, str]]]]:
    """
    Parse a call schedule spreadsheet into the backend storage format:

      { "YYYY-MM-DD": { "North Pod": [ {location, practitioner}, ... ], ... } }

    Supported formats: CSV and XLSX.

    Layout is auto-detected:
    - Header row: first row that has at least one cell that parses as a date.
    - Pod column: leftmost column that contains "North Pod", "Central Pod", or "South Pod" in data rows.
    - Date columns: columns to the right of the pod column whose header cells parse as dates.
    Rows above the header and columns to the left of the pod column are ignored.
    """
    if not file_bytes:
        raise ValueError("Uploaded file is empty. Please choose a non-empty schedule file")

    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(file_bytes)
    if name.endswith(".xlsx"):
        return _parse_xlsx(file_bytes)
    raise ValueError("Unsupported file type. Please upload a .csv or .xlsx file")


def _parse_csv(file_bytes: bytes) -> Dict[str, Dict[str, List[Dict[str, str]]]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("Spreadsheet is empty")

    header_row_idx = _find_header_row(rows)
    header = rows[header_row_idx]
    pod_col = _find_pod_column(rows, header_row_idx)
    date_columns = _find_date_columns(header, pod_col)
    if not date_columns:
        raise ValueError("No date columns found in header row")

    day_mapping = _empty_day_mapping_for_dates([d for _, d in date_columns])

    for row in rows[header_row_idx + 1 :]:
        if not row:
            continue
        pod = _normalize_pod(row[pod_col] if pod_col < len(row) else "")
        if not pod:
            continue
        for c, date_key in date_columns:
            if c >= len(row):
                continue
            val = str(row[c] or "").strip()
            for entry in _split_entries(val):
                location, practitioner = _parse_location_practitioner(entry)
                if not location and not practitioner:
                    continue
                day_mapping[date_key][pod].append(
                    {"location": location, "practitioner": practitioner}
                )

    return day_mapping


def _parse_xlsx(file_bytes: bytes) -> Dict[str, Dict[str, List[Dict[str, str]]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError("openpyxl is required to parse .xlsx uploads") from e

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    if ws.max_row == 0:
        raise ValueError("Spreadsheet is empty")

    # Build list of rows (same shape as CSV) so we can reuse header/pod detection.
    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row) if row else [])

    header_row_idx = _find_header_row(rows)
    header = rows[header_row_idx]
    pod_col = _find_pod_column(rows, header_row_idx)
    date_columns = _find_date_columns(header, pod_col)
    if not date_columns:
        raise ValueError("No date columns found in header row")

    day_mapping = _empty_day_mapping_for_dates([d for _, d in date_columns])

    for row in rows[header_row_idx + 1 :]:
        if not row:
            continue
        pod = _normalize_pod(row[pod_col] if pod_col < len(row) else "")
        if not pod:
            continue
        for c, date_key in date_columns:
            if c >= len(row):
                continue
            cell_val = row[c]
            if cell_val is None:
                continue
            val = str(cell_val).strip()
            for entry in _split_entries(val):
                location, practitioner = _parse_location_practitioner(entry)
                if not location and not practitioner:
                    continue
                day_mapping[date_key][pod].append(
                    {"location": location, "practitioner": practitioner}
                )

    return day_mapping

